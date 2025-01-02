#!/usr/bin/env python3

# Copyright (C) 2025 mald0rr
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import json
import argparse
from collections import Counter
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from rich.console import Console
from rich.table import Table
from rich import box
import networkx as nx
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font

class ScanAnalyzer:
    def __init__(self, json_file):
        with open(json_file, 'r') as f:
            self.data = json.load(f)
        self.console = Console()

    def generate_security_report(self, output_file=None):
        """Generate a detailed security report with findings and recommendations"""
        report = []
        
        # Header
        report.append("=" * 80)
        report.append(f"NETWORK SECURITY SCAN REPORT")
        report.append(f"Generated: {self.data['scan_time']}")
        report.append("=" * 80 + "\n")
        
        # 1. Executive Summary
        report.append("1. EXECUTIVE SUMMARY")
        report.append("-" * 80)
        total_hosts = len(self.data['hosts'])
        active_hosts = len([h for h in self.data['hosts'] if h['state'] == 'up'])
        total_open_ports = sum(len([p for p in host['ports'] if p['state'] == 'open']) 
                             for host in self.data['hosts'])
        total_vulns = sum(len(host.get('vulnerabilities', [])) for host in self.data['hosts'])
        
        report.append(f"Total hosts scanned: {total_hosts}")
        report.append(f"Active hosts: {active_hosts}")
        report.append(f"Total open ports: {total_open_ports}")
        report.append(f"Total vulnerabilities found: {total_vulns}\n")
        
        # 2. Network Infrastructure
        report.append("2. NETWORK INFRASTRUCTURE")
        report.append("-" * 80)
        for interface, info in self.data['network_info'].items():
            report.append(f"Interface: {interface}")
            report.append(f"  IP Address: {info['ip']}")
            report.append(f"  Netmask: {info['netmask']}")
            if 'network' in info:
                report.append(f"  Network: {info['network']}")
            report.append("")
        
        # 3. Host Analysis
        report.append("3. HOST ANALYSIS")
        report.append("-" * 80)
        for host in self.data['hosts']:
            report.append(f"\nHost: {host['ip']}")
            report.append(f"Status: {host['state']}")
            report.append(f"Hostname: {host['hostname'] if host['hostname'] else 'N/A'}")
            
            # OS Detection
            if host['os_matches']:
                os_match = host['os_matches'][0]
                report.append(f"Operating System: {os_match['name']} (Accuracy: {os_match['accuracy']}%)")
            
            # Open Ports
            open_ports = [p for p in host['ports'] if p['state'] == 'open']
            if open_ports:
                report.append("\nOpen Ports:")
                for port in open_ports:
                    report.append(f"  {port['port']}/{port['service']} - {port['product']} {port['version']}")
            
            report.append("")
        
        # 4. Vulnerability Assessment
        report.append("4. VULNERABILITY ASSESSMENT")
        report.append("-" * 80)
        vulns_found = False
        for host in self.data['hosts']:
            if host.get('vulnerabilities'):
                vulns_found = True
                report.append(f"\nHost: {host['ip']}")
                for vuln in host['vulnerabilities']:
                    report.append(f"\nPort {vuln['port']} - {vuln['type']}")
                    report.append("Details:")
                    report.append(f"{vuln['details']}")
                    report.append("-" * 40)
        
        if not vulns_found:
            report.append("\nNo vulnerabilities were detected with the current scan configuration.")
            
        # 5. Recommendations
        report.append("\n5. RECOMMENDATIONS")
        report.append("-" * 80)
        report.append("1. Review all open ports and disable unnecessary services")
        report.append("2. Ensure all services are running their latest stable versions")
        report.append("3. Implement firewall rules to restrict access to essential services only")
        report.append("4. Regular security scans should be performed to monitor changes")
        if total_vulns > 0:
            report.append("5. Address identified vulnerabilities based on severity and risk assessment")
        
        # Write to file or return as string
        report_text = '\n'.join(report)
        if output_file:
            with open(output_file, 'w') as f:
                f.write(report_text)
    def calculate_risk_scores(self):
        """Calculate risk scores for hosts based on vulnerabilities and exposure"""
        for host in self.data['hosts']:
            risk_score = 0
            exposures = []
            
            # Base exposure factors
            open_ports = len([p for p in host['ports'] if p['state'] == 'open'])
            critical_ports = len([p for p in host['ports'] 
                if p['state'] == 'open' and p['port'] in [22, 23, 3389, 445, 139, 135]])
            
            # Check for high-risk services
            for port in host['ports']:
                if port['state'] == 'open':
                    # Critical services check
                    if port['service'] in ['telnet', 'ftp']:
                        risk_score += 30
                        exposures.append(f"High-risk service: {port['service']} on port {port['port']}")
                    
                    # Outdated version detection (simplified)
                    if port.get('version') and any(x in port['version'].lower() 
                                                 for x in ['expired', 'outdated', 'old']):
                        risk_score += 20
                        exposures.append(f"Outdated service: {port['service']} ({port['version']})")

            # Vulnerability impact
            for vuln in host.get('vulnerabilities', []):
                if 'critical' in vuln['type'].lower():
                    risk_score += 40
                elif 'high' in vuln['type'].lower():
                    risk_score += 30
                elif 'medium' in vuln['type'].lower():
                    risk_score += 20
                elif 'low' in vuln['type'].lower():
                    risk_score += 10
            
            # Exposed critical ports
            if critical_ports > 0:
                risk_score += critical_ports * 15
                exposures.append(f"Critical ports exposed: {critical_ports}")
            
            # General exposure
            risk_score += open_ports * 5
            
            # Cap the risk score at 100
            risk_score = min(100, risk_score)
            
            # Add to host data
            host['risk_score'] = risk_score
            host['risk_exposures'] = exposures
            host['risk_level'] = 'Critical' if risk_score >= 80 else \
                               'High' if risk_score >= 60 else \
                               'Medium' if risk_score >= 40 else \
                               'Low'

    def check_compliance(self, compliance_standard='pci'):
        """Basic compliance checks for common standards"""
        compliance_issues = []
        
        for host in self.data['hosts']:
            host_issues = []
            
            # PCI DSS basic checks
            if compliance_standard.lower() == 'pci':
                # Check for prohibited services
                for port in host['ports']:
                    if port['state'] == 'open':
                        if port['service'] == 'telnet':
                            host_issues.append('Clear-text protocols (telnet) in use')
                        elif port['service'] == 'ftp':
                            host_issues.append('Insecure file transfer (FTP) enabled')
                
                # Check for exposed management interfaces
                mgmt_ports = [22, 23, 3389]
                exposed_mgmt = [p['port'] for p in host['ports'] 
                              if p['state'] == 'open' and p['port'] in mgmt_ports]
                if exposed_mgmt:
                    host_issues.append(f'Management ports exposed: {exposed_mgmt}')
            
            if host_issues:
                compliance_issues.append({
                    'ip': host['ip'],
                    'hostname': host['hostname'],
                    'issues': host_issues
                })
        
        return compliance_issues

    def detect_anomalies(self):
        """Detect potential security anomalies in the network"""
        anomalies = []
        
        # Collect standard port-service mappings
        standard_ports = {}
        service_ports = {}
        
        # First pass - collect normal mappings
        for host in self.data['hosts']:
            for port in host['ports']:
                if port['state'] == 'open':
                    service = port['service']
                    port_num = port['port']
                    
                    if service not in service_ports:
                        service_ports[service] = set()
                    service_ports[service].add(port_num)
                    
                    if port_num not in standard_ports:
                        standard_ports[port_num] = set()
                    standard_ports[port_num].add(service)
        
        # Second pass - detect anomalies
        for host in self.data['hosts']:
            host_anomalies = []
            
            for port in host['ports']:
                if port['state'] == 'open':
                    service = port['service']
                    port_num = port['port']
                    
                    # Unusual port for service
                    if service in service_ports and \
                       port_num not in service_ports[service] and \
                       len(service_ports[service]) > 0:
                        host_anomalies.append(
                            f"Unusual port {port_num} for service {service} " \
                            f"(normally on {service_ports[service]})")
                    
                    # Unusual service for port
                    if port_num in standard_ports and \
                       service not in standard_ports[port_num] and \
                       len(standard_ports[port_num]) > 0:
                        host_anomalies.append(
                            f"Unusual service {service} on port {port_num} " \
                            f"(normally {standard_ports[port_num]})")
            
            if host_anomalies:
                anomalies.append({
                    'ip': host['ip'],
                    'hostname': host['hostname'],
                    'anomalies': host_anomalies
                })
        
        return anomalies

    def display_service_summary(self):
        """Display a summary of running services across all hosts"""
        table = Table(title="Service Summary", box=box.ROUNDED)
        table.add_column("Service", style="cyan")
        table.add_column("Count", style="magenta")
        table.add_column("Ports", style="green")
        
        services = {}
        for host in self.data['hosts']:
            for port in host['ports']:
                if port['state'] == 'open':
                    service = port['service']
                    if service not in services:
                        services[service] = {'count': 0, 'ports': set()}
                    services[service]['count'] += 1
                    services[service]['ports'].add(str(port['port']))
        
        for service, info in sorted(services.items(), key=lambda x: x[1]['count'], reverse=True):
            table.add_row(
                service,
                str(info['count']),
                ", ".join(sorted(info['ports'], key=int))
            )
        
        self.console.print(table)

    def create_network_graph(self, output_file='network_graph.png'):
        """Create a visual representation of the network"""
        G = nx.Graph()
        
        # Add all hosts and their services
        host_status = {}  # Track which hosts have services
        service_counts = {}  # Track services for grouping
        
        for host in self.data['hosts']:
            ip = host['ip']
            open_ports = [p for p in host['ports'] if p['state'] == 'open']
            
            # Add all hosts
            G.add_node(ip, type='host', has_services=bool(open_ports))
            host_status[ip] = bool(open_ports)
            
            # Add service nodes and connect them to hosts
            if open_ports:
                for port in open_ports:
                    service_name = f"{port['service']}\n({port['port']})"
                    if service_name not in service_counts:
                        service_counts[service_name] = len(service_counts)
                    G.add_node(service_name, type='service', 
                             service_group=service_counts[service_name])
                    G.add_edge(ip, service_name)
        
        # Create the visualization with a larger figure
        plt.figure(figsize=(20, 15))
        
        # Use a circular layout for hosts and group services by type
        host_nodes = [node for node, attr in G.nodes(data=True) if attr['type'] == 'host']
        service_nodes = [node for node, attr in G.nodes(data=True) if attr['type'] == 'service']
        
        # Position calculation
        # Position hosts in a circle
        host_pos = nx.circular_layout(G.subgraph(host_nodes), scale=2)
        
        # Position services in groups based on their type
        service_pos = {}
        for service in service_nodes:
            # Get the connected host positions
            connected_hosts = list(G.neighbors(service))
            # Calculate average position of connected hosts
            avg_x = np.mean([host_pos[host][0] for host in connected_hosts])
            avg_y = np.mean([host_pos[host][1] for host in connected_hosts])
            # Add some variation based on service group
            group = G.nodes[service]['service_group']
            angle = 2 * np.pi * group / len(service_counts) if service_counts else 0
            service_pos[service] = [
                avg_x + 0.5 * np.cos(angle),
                avg_y + 0.5 * np.sin(angle)
            ]
        
        # Combine positions
        pos = {**host_pos, **service_pos}
        
        # Draw hosts with different colors based on whether they have services
        hosts_with_services = [node for node in host_nodes if host_status[node]]
        hosts_without_services = [node for node in host_nodes if not host_status[node]]
        
        # Draw hosts with services
        if hosts_with_services:
            nx.draw_networkx_nodes(G, pos, nodelist=hosts_with_services, 
                                 node_color='lightblue',
                                 node_size=3000, alpha=0.8)
        
        # Draw hosts without services (in a different color)
        if hosts_without_services:
            nx.draw_networkx_nodes(G, pos, nodelist=hosts_without_services, 
                                 node_color='lightgray',
                                 node_size=3000, alpha=0.8)
        
        # Draw services
        for service in service_nodes:
            nx.draw_networkx_nodes(G, pos, 
                                 nodelist=[service],
                                 node_color=['lightgreen'],
                                 node_size=2000, alpha=0.8)
        
        # Draw edges with curved lines
        nx.draw_networkx_edges(G, pos, alpha=0.5, 
                             edge_color='gray',
                             connectionstyle="arc3,rad=0.2")
        
        # Draw labels with better spacing
        nx.draw_networkx_labels(G, pos, font_size=10,
                              bbox=dict(facecolor='white', 
                                      edgecolor='none', 
                                      alpha=0.7,
                                      pad=4))
        
        # Add a legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor='lightblue', alpha=0.8, label='Hosts with services'),
            Patch(facecolor='lightgray', alpha=0.8, label='Hosts without services'),
            Patch(facecolor='lightgreen', alpha=0.8, label='Services')
        ]
        plt.legend(handles=legend_elements, loc='upper right')
        
        plt.title("Network Service Map")
        plt.axis('off')
        plt.savefig(output_file, bbox_inches='tight', dpi=300)
        plt.close()

    def export_to_excel(self, output_file='network_scan.xlsx'):
        """Export scan results to Excel format with enhanced vulnerability reporting"""
        with pd.ExcelWriter(output_file) as writer:
            # Host summary with vulnerability counts
            host_data = []
            for host in self.data['hosts']:
                os_name = host['os_matches'][0]['name'] if host['os_matches'] else 'Unknown'
                open_ports = len([p for p in host['ports'] if p['state'] == 'open'])
                vuln_count = len(host.get('vulnerabilities', []))
                
                host_data.append({
                    'IP': host['ip'],
                    'Hostname': host['hostname'],
                    'OS': os_name,
                    'Open Ports': open_ports,
                    'Vulnerabilities Found': vuln_count,
                    'State': host['state']
                })
            
            df_hosts = pd.DataFrame(host_data)
            
            # Add total row
            totals = pd.Series({
                'IP': 'TOTAL',
                'Hostname': f'{len(host_data)} hosts',
                'Open Ports': df_hosts['Open Ports'].sum(),
                'Vulnerabilities Found': df_hosts['Vulnerabilities Found'].sum()
            })
            df_hosts = pd.concat([df_hosts, pd.DataFrame([totals])], ignore_index=True)
            
            # Write hosts summary
            df_hosts.to_excel(writer, sheet_name='Hosts Summary', index=False)
            
            # Format the totals row using openpyxl
            workbook = writer.book
            worksheet = workbook['Hosts Summary']
            gray_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            bold_font = Font(bold=True)
            
            last_row = len(df_hosts)
            for cell in worksheet[last_row]:
                cell.fill = gray_fill
                cell.font = bold_font
            
            # Port details with vulnerability indicators
            port_data = []
            for host in self.data['hosts']:
                host_vulns = {v['port']: v['type'] for v in host.get('vulnerabilities', [])}
                for port in host['ports']:
                    port_data.append({
                        'IP': host['ip'],
                        'Port': port['port'],
                        'State': port['state'],
                        'Service': port['service'],
                        'Product': port['product'],
                        'Version': port['version'],
                        'Has Vulnerability': 'Yes' if port['port'] in host_vulns else 'No',
                        'Vulnerability Type': host_vulns.get(port['port'], 'N/A')
                    })
            
            df_ports = pd.DataFrame(port_data)
            df_ports.to_excel(writer, sheet_name='Port Details', index=False)
            
            # Vulnerabilities detail sheet
            vuln_data = []
            for host in self.data['hosts']:
                for vuln in host.get('vulnerabilities', []):
                    vuln_data.append({
                        'IP': host['ip'],
                        'Port': vuln['port'],
                        'Type': vuln['type'],
                        'Details': vuln['details']
                    })
            
            if vuln_data:
                df_vulns = pd.DataFrame(vuln_data)
                df_vulns.to_excel(writer, sheet_name='Vulnerabilities', index=False)
            
            # Add a vulnerability summary sheet
            if vuln_data:
                vuln_summary = []
                # Group vulnerabilities by type
                vuln_types = {}
                for vuln in vuln_data:
                    vuln_type = vuln['Type']
                    if vuln_type not in vuln_types:
                        vuln_types[vuln_type] = 0
                    vuln_types[vuln_type] += 1
                
                for vuln_type, count in vuln_types.items():
                    vuln_summary.append({
                        'Vulnerability Type': vuln_type,
                        'Count': count,
                        'Affected Hosts': len(set(v['IP'] for v in vuln_data if v['Type'] == vuln_type))
                    })
                
                df_vuln_summary = pd.DataFrame(vuln_summary)
                df_vuln_summary = df_vuln_summary.sort_values('Count', ascending=False)
                df_vuln_summary.to_excel(writer, sheet_name='Vulnerability Summary', index=False)
                
                # Format vulnerability summary sheet using openpyxl
                worksheet = workbook['Vulnerability Summary']
                green_fill = PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid')
                bold_font = Font(bold=True)
                
                # Format header row
                for cell in worksheet[1]:
                    cell.fill = green_fill
                    cell.font = bold_font
                
                # Adjust column widths
                for col, width in enumerate([30, 10, 15], start=1):
                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

def main():
    parser = argparse.ArgumentParser(description='Analyze network scan results')
    parser.add_argument('json_file', help='Input JSON file from network scan')
    parser.add_argument('--report', help='Generate security report to specified file')
    parser.add_argument('--excel', help='Export to Excel file')
    parser.add_argument('--graph', help='Create network graph visualization')
    parser.add_argument('--services', action='store_true', help='Display service summary')
    parser.add_argument('--mitigation', help='Generate mitigation plan report')
    parser.add_argument('--segmentation', help='Generate network segmentation analysis report')
    
    args = parser.parse_args()
    
    analyzer = ScanAnalyzer(args.json_file)
    
    # Calculate risk scores for all outputs
    analyzer.calculate_risk_scores()
    
    if args.report:
        analyzer.generate_security_report(args.report)
        print(f"Security report generated: {args.report}")
    
    if args.excel:
        analyzer.export_to_excel(args.excel)
        print(f"Excel report generated: {args.excel}")
    
    if args.graph:
        analyzer.create_network_graph(args.graph)
        print(f"Network graph generated: {args.graph}")
    
    if args.services:
        analyzer.display_service_summary()
        
    if args.mitigation:
        mitigation_plan = analyzer.generate_mitigation_plan()
        with open(args.mitigation, 'w') as f:
            f.write("NETWORK SECURITY MITIGATION PLAN\n")
            f.write("=" * 50 + "\n\n")
            
            for host in mitigation_plan:
                f.write(f"Host: {host['ip']} ({host['hostname']})\n")
                f.write(f"Risk Score: {host['risk_score']}\n")
                f.write("-" * 50 + "\n")
                
                for mitigation in host['mitigations']:
                    f.write(f"\nIssue: {mitigation['target']}\n")
                    f.write(f"Risk Level: {mitigation['risk']}\n")
                    f.write(f"Recommended Fix: {mitigation['fix']}\n")
                    f.write(f"Effort Level: {mitigation['effort']}\n")
                    f.write(f"Security Benefit: {mitigation['benefit']}\n")
                    f.write("\nImplementation Steps:\n")
                    for i, step in enumerate(mitigation['steps'], 1):
                        f.write(f"{i}. {step}\n")
                    f.write("\n")
                f.write("\n")
        print(f"Mitigation plan generated: {args.mitigation}")
    
    if args.segmentation:
        segmentation_analysis = analyzer.analyze_network_segmentation()
        with open(args.segmentation, 'w') as f:
            f.write("NETWORK SEGMENTATION ANALYSIS\n")
            f.write("=" * 50 + "\n\n")
            
            # Global recommendations
            if segmentation_analysis['recommendations']:
                f.write("Global Recommendations:\n")
                f.write("-" * 30 + "\n")
                for rec in segmentation_analysis['recommendations']:
                    f.write(f"* {rec}\n")
                f.write("\n")
            
            # Segment analysis
            f.write("Segment Analysis:\n")
            f.write("-" * 30 + "\n")
            for segment in segmentation_analysis['segments']:
                f.write(f"\nNetwork: {segment['network']}\n")
                f.write(f"Hosts: {segment['host_count']}\n")
                f.write(f"Risk Levels: {', '.join(segment['risk_levels'])}\n")
                f.write(f"Services: {', '.join(segment['services'])}\n")
                
                if segment['issues']:
                    f.write("\nIssues:\n")
                    for issue in segment['issues']:
                        f.write(f"! {issue}\n")
                
                if segment['recommendations']:
                    f.write("\nRecommendations:\n")
                    for rec in segment['recommendations']:
                        f.write(f"* {rec}\n")
                f.write("\n")
        print(f"Segmentation analysis generated: {args.segmentation}")


if __name__ == "__main__":
    main()
